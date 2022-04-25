import pickle
#import subprocess

import openpyxl.reader.excel
from openpyxl import Workbook
import gi
import os

gi.require_version('Gtk', '3.0')
from gi.repository import Gtk as gtkr
from gi.overrides import Gtk as gtk


class Book:
    def __init__(self, title, author, status, tags, path, remarks):
        self.title = title
        self.author = author
        self.status = status
        self.tags = tags
        self.path = path
        self.remarks = remarks


# #########Initialize pickle file###########
# books = []
# pickle_out = open('list.pickle', 'wb')
# pickle.dump(books, pickle_out)
# pickle_out.close()
# ##########################################

class Main:
    def __init__(self):
        gladeFile = 'unsaved.glade'
        self.builder = gtk.Builder()
        self.builder.add_from_file(gladeFile)

        self.builder.connect_signals(
            {'add_btn_clicked': self.add_btn_clicked, 'reader_btn_clicked': self.reader_btn_clicked,
             'remarks_btn_clicked': self.remarks_btn_clicked,
             'on_treeview_row_activated': self.on_treeview_row_activated, 'setSelected': self.setSelected,
             'button_press': self.button_press, 'remove': self.remove, 'load_list': self.load_list,
             'save_list': self.save_list, 'open': self.open, 'edit': self.edit,
             'save_btn': self.save_btn, 'path_btn_clicked': self.path_btn_clicked})

        self.treeview = self.builder.get_object('treeview')
        self.liststore = self.builder.get_object('liststore2')

        self.textbuffer = self.builder.get_object('textbuffer')
        self.GtkMenu = self.builder.get_object('GtkMenu')
        self.deletebtn = self.builder.get_object('deletebtn')
        self.editbtn = self.builder.get_object('editbtn')

        # self.editbtn.connect('button-press-event', self.test)

        print('init start')
        self.global_row = 0

        self.reload_liststore()

        window = self.builder.get_object('Mainwindow')
        window.connect('delete-event', gtk.main_quit)
        window.show()

    def open(self, widget, b):

        model = widget.get_model()
        x = model[self.global_row][0]
        pickle_in = open('list.pickle', 'rb')
        pickle_list = pickle.load(pickle_in)

        for i, o in enumerate(pickle_list):
            if o.title == x:
                #subprocess.Popen(['/home/vito/Desktop/Gladetest/List of books.xlsx'], shell=True)

                # os.system('file:///home/vito/Desktop/python-gtk-3-tutorial.pdf')
                # os.startfile('file:///home/vito/Desktop/python-gtk-3-tutorial.pdf')

                break

        print('edit')

    def save_list(self, a, b):
        dialog = self.builder.get_object('GtkFileChooserSave')


        dialog.set_do_overwrite_confirmation(True)


        response = dialog.run()

        help(dialog)

        if response == gtkr.ResponseType.OK:

            u = dialog.get_uri()
            f = dialog.get_filename()

            print(u)
            print(f)

            wb = Workbook()
            ws = wb.active
            pickle_in = open('list.pickle', 'rb')
            pickle_list = pickle.load(pickle_in)
            ws.append(['№', 'Title', 'Author', 'Status', 'Tags', 'Path', 'Remarks'])

            for i, o in enumerate(pickle_list):
                ws.append([i,
                           o.title,
                           o.author,
                           o.status,
                           o.tags,
                           o.path,
                           o.remarks])
                wb.save(f)

            dialog.hide()
        elif response == gtkr.ResponseType.CANCEL:
            print('cancel')
            dialog.hide()
        dialog.hide()
        print('save list')

        # if o.title == x:
        #    pass
        # ws['A2'] = 42
        # ws.append([1, 2, 3])
        # ws['A2'] = 55

    def load_list(self, a, b):
        dialog = self.builder.get_object('GtkFileChooser')


        response = dialog.run()
        if response == gtkr.ResponseType.OK:
            u = dialog.get_filename()
            print(u)
            wb = openpyxl.reader.excel.load_workbook(filename=u)
            sheet = wb.active
            pickle_in = open('list.pickle', 'rb')
            pickle_list = pickle.load(pickle_in)
            pickle_list.clear()
            for i in range(2, 1000):
                if sheet['A' + str(i)].value is not None:

                    no = sheet['A' + str(i)].value
                    n = sheet['B' + str(i)].value
                    a = sheet['C' + str(i)].value
                    s = sheet['D' + str(i)].value
                    t = sheet['E' + str(i)].value
                    p = sheet['F' + str(i)].value
                    r = sheet['G' + str(i)].value

                    book = Book(title=n if n != '' else 'Undefined',
                                author=a if a != '' else 'Undefined',
                                status=s if s != '' else 'Undefined',
                                tags=t if t != '' else 'Undefined',
                                path=p if p != '' else 'Undefined',
                                remarks=r if r != '' else 'Undefined')
                    pickle_list.append(book)

                else:
                    break

            pickle_out = open('list.pickle', 'wb')
            pickle.dump(pickle_list, pickle_out)
            pickle_out.close()

            self.reload_liststore()

            dialog.hide()
        elif response == gtkr.ResponseType.CANCEL:
            print('cancel')
            dialog.hide()
        dialog.hide()
        print('load')

    def path_btn_clicked(self, a):
        dialog = self.builder.get_object('GtkFileChooser')
        response = dialog.run()
        if response == gtkr.ResponseType.OK:
            u = dialog.get_uri()
            print(u)
            self.builder.get_object('Path_input').set_text(u)
            dialog.hide()
        elif response == gtkr.ResponseType.CANCEL:
            print('cancel')
            dialog.hide()
        dialog.hide()
        print('path')

    def button_press(self, widget, event):

        if event.button == 3:
            self.GtkMenu.popup(None, None, None, None, event.button, event.time)

    def save_btn(self, widget):

        model = widget.get_model()
        x = model[self.global_row][0]
        start = self.textbuffer.get_start_iter()
        end = self.textbuffer.get_end_iter()

        pickle_in = open('list.pickle', 'rb')
        pickle_list = pickle.load(pickle_in)

        for i, o in enumerate(pickle_list):
            if o.title == x:
                o.remarks = self.textbuffer.get_text(start, end, self)
                break
        pickle_out = open('list.pickle', 'wb')
        pickle.dump(pickle_list, pickle_out)
        pickle_out.close()

        # self.reload_liststore()

    def reload_liststore(self):
        self.liststore.clear()
        print('reload')
        pickle_in = open('list.pickle', 'rb')
        new_loaded_list = pickle.load(pickle_in)
        if new_loaded_list:
            for i in new_loaded_list:
                print(i.title, i.author, i.status, i.tags, i.path, i.remarks)
                list = [(i.title), (i.author), (i.status), (i.tags), (i.path)]
                self.liststore.append(list)
        else:
            print('No books yet')
            list = [('No books yet'), (''), (''), (''), ('')]
            self.liststore.append(list)

    def on_treeview_row_activated(self, widget, row, col_view):
        self.global_row = row
        print('global row is ' + str(self.global_row))

        model = widget.get_model()

        x = model[self.global_row][0]

        pickle_in = open('list.pickle', 'rb')
        pickle_list = pickle.load(pickle_in)

        for i, o in enumerate(pickle_list):
            if o.title == x:
                print(o.remarks)
                self.textbuffer.set_text(o.remarks)
                break

    def add_btn_clicked(self, widget):
        dialog = self.builder.get_object('GtkPopover1')
        response = dialog.run()
        if response == gtkr.ResponseType.OK:
            n = self.builder.get_object('Title_input').get_text()
            a = self.builder.get_object('Author_input').get_text()
            s = self.builder.get_object('Status_input').get_text()
            t = self.builder.get_object('Tags_input').get_text()
            p = self.builder.get_object('Path_input').get_text()
            self.builder.get_object('Title_input').set_text('')
            self.builder.get_object('Author_input').set_text('')
            self.builder.get_object('Status_input').set_text('')
            self.builder.get_object('Tags_input').set_text('')
            self.builder.get_object('Path_input').set_text('')

            dialog.hide()

            # p = ''
            r = ''

            book = Book(title=n if n != '' else 'Undefined',
                        author=a if a != '' else 'Undefined',
                        status=s if s != '' else 'Undefined',
                        tags=t if t != '' else 'Undefined',
                        path=p if p != '' else 'Undefined',
                        remarks=r if r != '' else 'Undefined')

            print(book.title, book.author, book.status, book.remarks, book.path)

            pickle_in = open('list.pickle', 'rb')
            pickle_list = pickle.load(pickle_in)
            pickle_list.append(book)

            pickle_out = open('list.pickle', 'wb')
            pickle.dump(pickle_list, pickle_out)
            pickle_out.close()

            self.reload_liststore()

            print('add_btn_clicked')

        elif response == gtkr.ResponseType.CANCEL:
            dialog.hide()
        dialog.hide()

    def reader_btn_clicked(self, a):
        self.GtkRevealer2 = self.builder.get_object('GtkRevealer2')
        f = self.GtkRevealer2.get_reveal_child()
        print(f)
        self.GtkRevealer2.set_reveal_child(True if f is False else False)

    def remarks_btn_clicked(self, a):
        print('remarks_btn_clicked')
        # print(a.row_activated)
        self.GtkRevealer = self.builder.get_object('GtkRevealer')
        r = self.GtkRevealer.get_reveal_child()
        print(r)
        self.GtkRevealer.set_reveal_child(True if r is False else False)

    def remove(self, widget, b):
        model = widget.get_model()
        print('removing:')
        print(model[self.global_row][0])
        x = model[self.global_row][0]

        pickle_in = open('list.pickle', 'rb')
        pickle_list = pickle.load(pickle_in)

        for i, o in enumerate(pickle_list):
            if o.title == x:
                del pickle_list[i]
                break

        pickle_out = open('list.pickle', 'wb')
        pickle.dump(pickle_list, pickle_out)
        pickle_out.close()
        self.reload_liststore()
        return

    def edit(self, widget, b):
        model = widget.get_model()
        print('edit:')
        print(model[self.global_row][0])
        x = model[self.global_row][0]

        pickle_in = open('list.pickle', 'rb')
        pickle_list = pickle.load(pickle_in)

        for i, o in enumerate(pickle_list):
            if o.title == x:

                self.builder.get_object('Title_input').set_text(o.title)
                self.builder.get_object('Author_input').set_text(o.author)
                self.builder.get_object('Status_input').set_text(o.status)
                self.builder.get_object('Tags_input').set_text(o.tags)
                self.builder.get_object('Path_input').set_text(o.path)
                dialog = self.builder.get_object('GtkPopover1')
                response = dialog.run()

                if response == gtkr.ResponseType.OK:
                    n = self.builder.get_object('Title_input').get_text()
                    a = self.builder.get_object('Author_input').get_text()
                    s = self.builder.get_object('Status_input').get_text()
                    t = self.builder.get_object('Tags_input').get_text()
                    self.builder.get_object('Title_input').set_text('')
                    self.builder.get_object('Author_input').set_text('')
                    self.builder.get_object('Status_input').set_text('')
                    self.builder.get_object('Tags_input').set_text('')
                    dialog.hide()
                    p = o.path
                    r = o.remarks

                    book = Book(title=n if n != '' else 'Undefined',
                                author=a if a != '' else 'Undefined',
                                status=s if s != '' else 'Undefined',
                                tags=t if t != '' else 'Undefined',
                                path=p if p != '' else 'Undefined',
                                remarks=r if r != '' else 'Undefined')

                    print(book.title, book.author, book.status, book.remarks, book.path)
                    self.remove(widget, b)

                    pickle_in = open('list.pickle', 'rb')
                    pickle_list = pickle.load(pickle_in)
                    pickle_list.append(book)

                    pickle_out = open('list.pickle', 'wb')
                    pickle.dump(pickle_list, pickle_out)
                    pickle_out.close()
                    self.reload_liststore()

                    print('edit_btn_clicked')

                elif response == gtkr.ResponseType.CANCEL:
                    dialog.hide()
                dialog.hide()

    # model = widget.get_widget()
    # row_list = []
    # for i in range(model.get_n_columns()):
    #     # print(model[row][i])
    #     row_list.append(model[row][i])

    def setSelected(self, selection):
        pass


if __name__ == '__main__':
    main = Main()
    gtk.main()
